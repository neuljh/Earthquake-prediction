import openpyxl

# 创建一个新的Excel工作簿
workbook = openpyxl.Workbook()
# 选择默认的活动工作表
sheet = workbook.active
# 定义要插入数据的列属性
column_attributes = {
    'A': 'Name',
    'B': 'Age',
    'C': 'City'
}
# 在第一行插入列属性
for column, attribute in column_attributes.items():
    sheet[column + '1'] = attribute

# 保存Excel文件
workbook.save('data.xlsx')
workbook = openpyxl.load_workbook('data.xlsx')
# 选择默认的活动工作表
sheet = workbook.active

# 定义要插入的数据
data = [
    ['John', 25, 'New York'],
    ['Emma', 30, 'London'],
    ['Michael', 35, 'Sydney']
]

# 获取最后一行的索引
last_row = sheet.max_row + 1

# 插入数据到指定列
for row in data:
    for column, value in zip(column_attributes.keys(), row):
        sheet[column + str(last_row)] = value
    last_row += 1

# 保存Excel文件
workbook.save('data.xlsx')