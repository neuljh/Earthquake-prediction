import pandas as pd
import numpy as np
from scipy.fft import fft, fftfreq
import matplotlib.pyplot as plt
import os
import pandas as pd
from scipy.signal import correlate
import openpyxl

# 输入和输出目录
input_dir = 'outputdata'
output_dir = 'outputdatafft'

# 获取输入目录下的子文件夹列表
subfolders = os.listdir(input_dir)

# 循环遍历每个子文件夹
for subfolder in subfolders:
    # 构建输入和输出子文件夹路径
    input_subfolder = os.path.join(input_dir, subfolder)
    output_subfolder = os.path.join(output_dir, subfolder)

    # 创建输出子文件夹
    os.makedirs(output_subfolder, exist_ok=True)

    # 获取输入子文件夹中的文件列表
    files = [file for file in os.listdir(input_subfolder) if file.endswith('.xlsx')]

    # 循环遍历每个文件
    for file in files:
        # 构建输入和输出文件路径
        input_file = os.path.join(input_subfolder, file)
        output_file = os.path.join(output_subfolder, file.replace('.xlsx', ''))
        os.makedirs(output_file)

        print('input_file: '+input_file)
        print('output_file: ' + output_file)

        # 读取Excel数据
        df = pd.read_excel(input_file, header=None)

        # 创建一个新的Excel工作簿
        workbook = openpyxl.Workbook()
        # 选择默认的活动工作表
        sheet = workbook.active
        # 定义要插入数据的列属性
        column_attributes = {
            'A': 'Mean',
            'B': 'Standard Deviation',
            'C': 'Peak Value',
            'D': 'Duration',
            'E': 'Peak Frequency',
            'F': 'Bandwidth',
            'G': 'Autocorrelation Peak',
            'H': 'Autocorrelation Lag',
            'I': 'Entropy',
            'J': 'Energy'
        }
        # 在第一行插入列属性
        for column, attribute in column_attributes.items():
            sheet[column + '1'] = attribute

        # 保存Excel文件
        workbook.save(output_file+'.xlsx')
        workbook = openpyxl.load_workbook(output_file+'.xlsx')
        # 选择默认的活动工作表
        sheet = workbook.active
        # 定义要插入的数据
        datas_list = []

        # 遍历每一行数据
        for index, row in df.iterrows():
            data = row.values  # 获取一行数据

            # 计算平均值
            mean = np.mean(row)
            # 计算标准差
            std = np.std(row)

            # 频谱分析
            N = len(data)  # 数据长度
            T = 1.0 / N  # 采样间隔
            yf = fft(data)  # 快速傅里叶变换
            xf = fftfreq(N, T)[:N // 2]  # 频率

            # 计算频谱峰值频率
            peak_freq = xf[np.argmax(np.abs(yf[:N // 2]))]

            # 计算频谱带宽
            threshold = 0.5  # 能量阈值
            spectral_energy = np.abs(yf[:N // 2]) ** 2
            total_energy = np.sum(spectral_energy)
            cum_energy = np.cumsum(spectral_energy) / total_energy
            bandwidth = xf[np.where(cum_energy >= threshold)[0][0]] - xf[np.where(cum_energy >= threshold)[0][-1]]

            # 计算峰值
            peak_value = row.max()
            # 计算持续时间
            duration = len(row)

            # 计算自相关函数
            autocorr = correlate(row, row, mode='full')
            # 提取自相关峰值和自相关时延
            autocorr_peak = np.max(autocorr)
            autocorr_lag = np.argmax(autocorr)

            # 计算熵值
            entropy = -np.sum(row * np.log2(np.abs(row)))
            # 计算能量
            energy = np.sum(row ** 2)

            # 打印结果
            print(f"Row {index + 1}:Mean = {mean}, Standard Deviation = {std} Entropy = {entropy}, Energy = {energy}")
            print(f"Row {index + 1}: Peak Frequency = {peak_freq}, Bandwidth = {bandwidth} Peak Value = {peak_value}, Duration = {duration} Autocorrelation Peak = {autocorr_peak}, Autocorrelation Lag = {autocorr_lag} ")

            # 保存频谱图
            file_name = f'spectrum_row{index + 1}.png'
            file_path = os.path.join(output_file, file_name)
            # 绘制频谱图
            plt.figure()
            plt.plot(xf, 2.0 / N * np.abs(yf[:N // 2]))
            plt.xlabel('Frequency')
            plt.ylabel('Amplitude')
            plt.title(f'Spectrum - Row {index + 1}')
            plt.savefig(file_path)
            plt.show()

            # 添加数据
            data_list=[]
            data_list.append(mean)
            data_list.append(std)
            data_list.append(peak_value)
            data_list.append(duration)
            data_list.append(peak_freq)
            data_list.append(bandwidth)
            data_list.append(autocorr_peak)
            data_list.append(autocorr_lag)
            data_list.append(entropy)
            data_list.append(energy)
            datas_list.append(data_list)
        # 获取最后一行的索引
        last_row = sheet.max_row + 1
        # 插入数据到指定列
        for row in datas_list:
            for column, value in zip(column_attributes.keys(), row):
                sheet[column + str(last_row)] = value
            last_row += 1
        # 保存Excel文件
        workbook.save(output_file+'.xlsx')






